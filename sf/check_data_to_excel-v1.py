#!/usr/bin/env python3 
# -*- coding:utf-8 _*-
""" 
@author:fjg
@license: Apache Licence 
@file: check_data_to_excel-v1.py
@time: 2021/02/22
@contact: fujiangong.fujg@bytedance.com
@site:  
@software: PyCharm 
"""
import argparse
import re
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


def existing_path(value: str) -> str:
    """
    判断地址是否存在
    :param value:
    :return:
    """
    path = Path(value)
    if not path.exists():
        raise ValueError(f"Path does not exist: {path}")
    return value


def check_output_file(value: str):
    path = Path(value)
    if path.exists():
        path.unlink()


def get_parse() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument("data_dir", type=existing_path,
                        help="Check results directory. it must contain： cluster check result、node check result \
                        and k8s core component check result")
    parser.add_argument("cluster_name", type=str, help="Generated Excel file name.")
    return parser


def set_ranger_cell_values(ws_sheet, ranger_cell, value_tuple):
    min_row = ws_sheet[ranger_cell][0][0].row
    min_column = ws_sheet[ranger_cell][0][0].column
    for a in ws_sheet[ranger_cell]:
        for b in a:
            try:
                b.value = value_tuple[b.row - min_row][b.column - min_column]
            except AttributeError:
                continue


def set_column_width(wb_sheet: Workbook.worksheets, start_column: str, end_column: str, width: int):
    for i in range(ord(start_column), ord(end_column) + 1):
        wb_sheet.column_dimensions[chr(i)].width = width


def set_row_height(wb_sheet, start_row, end_row, height):
    for i in range(start_row, end_row + 1):
        wb_sheet.row_dimensions[i].height = height


def set_range_styles(ws_sheet: Workbook.worksheets, ranger_cell: str, font: Font = None, fill: PatternFill = None,
                     alignment: Alignment = None, border: Border = None):
    for a in ws_sheet[ranger_cell]:
        for b in a:
            if font is not None:
                b.font = font
            if fill is not None:
                b.fill = fill
            if alignment is not None:
                b.alignment = alignment
            if border is not None:
                b.border = border


def get_suffix_file(path: str, suffix: str) -> list:
    file_list = []
    path = path.rstrip("/")
    file_dir = Path(path).iterdir()
    for it in file_dir:
        if it.is_file():
            if it.suffix == suffix:
                file_list.append("{0}/{1}".format(path, it.name))
    return file_list


def check_file_from_what(file_list: list) -> dict:
    check_data_dict = dict()
    for file in file_list:
        with open(file, 'r') as check_file:
            file_data = check_file.read()
            if re.search("diskUsage", file_data):
                check_data_dict["node"] = file
            elif re.search("pod_status", file_data):
                check_data_dict["cluster"] = file
            elif re.search("apiserver_process", file_data):
                check_data_dict["component"] = file
    return check_data_dict


def format_txt_to_json(path: str) -> dict:
    cluster_info = dict()
    last_host_ip = ""
    last_check_point = ""
    with open(path, 'r', encoding='utf-8') as check_file:
        for line in check_file.readlines():
            if line.strip() in ["", "^M"]:
                continue
            ip_count = re.match(r'^(?:25[0-5]\.|2[0-4]\d\.|[01]?\d\d?\.){3}(?:25[0-5]|2[0-4]\d|[01]?\d\d?)',
                                line.strip())
            if ip_count:
                last_host_ip = ip_count.group(0)
                cluster_info[last_host_ip] = {}
                continue
            try:
                check_dict = json.loads(line.strip())
                last_check_point = check_dict["check_point"]
                del check_dict["check_point"]
                if last_check_point in cluster_info[last_host_ip].keys():
                    cluster_info[last_host_ip][last_check_point].append(check_dict)
                else:
                    cluster_info[last_host_ip][last_check_point] = []
                    cluster_info[last_host_ip][last_check_point].append(check_dict)
            except json.decoder.JSONDecodeError:
                if "others" in cluster_info[last_host_ip][last_check_point][-1].keys():
                    cluster_info[last_host_ip][last_check_point][-1]["others"].append(line)
                else:
                    cluster_info[last_host_ip][last_check_point][-1]["others"] = []
                    cluster_info[last_host_ip][last_check_point][-1]["others"].append(line)
    return cluster_info


def set_cell(sheet: Workbook.worksheets, row: int, column: int, values: str, status: str = "info"):
    cell = sheet.cell(row=row, column=column)
    alignment = Alignment(horizontal="center", vertical="center")
    if status == "info":
        fill = PatternFill()
    elif status == "warning":
        fill = PatternFill(patternType="solid", start_color="FFFF00")
    elif status == "error":
        fill = PatternFill(patternType="solid", start_color="FF0000")
    else:
        fill = PatternFill()
    cell.value = values
    cell.fill = fill
    cell.alignment = alignment


def cluster_json_to_excel(cluster_dict: dict, sheet: Workbook.worksheets):
    sheet.cell(row=2, column=1, value=args.cluster_name)
    for ip in cluster_dict.keys():
        items = cluster_dict[ip]
        if "node_status" in items.keys():
            key_data = items["node_status"]
            alert_status = key_data[0]["alert_status"]
            ready = key_data[0]["check_data"]["readyNodeNum"]
            not_ready = key_data[0]["check_data"]["notReadyNodeNum"]
            set_cell(sheet, 3, 4, ready)
            set_cell(sheet, 4, 4, not_ready, alert_status)
        else:
            set_cell(sheet, 3, 4, "无结果", "error")
            set_cell(sheet, 4, 4, "无结果", "error")
        if "pod_status" in items.keys():
            info_status = ["Running", "Completed"]
            error_status = ["CrashLoopBackOff", "ImagePullBackOff", "ContainerCreating", "Terminating", "ERROR",
                            "restart"]
            key_data = items["pod_status"]
            for pods_status in key_data:
                status = pods_status["check_data"]["status"]
                nums = pods_status["check_data"]["num"]
                alert_status = pods_status["alert_status"]
                if status in info_status:
                    if status == "Running":
                        set_cell(sheet, 6, 4, nums)
                    elif status == "Completed":
                        set_cell(sheet, 7, 4, nums)
                if status in error_status:
                    other_info = "".join(pods_status["others"])
                    if status == "CrashLoopBackOff":
                        set_cell(sheet, 8, 4, nums, alert_status)
                        set_cell(sheet, 8, 5, other_info, alert_status)
                    elif status == "ImagePullBackOff":
                        set_cell(sheet, 9, 4, nums, alert_status)
                        set_cell(sheet, 9, 5, other_info, alert_status)
                    elif status == "ContainerCreating":
                        set_cell(sheet, 10, 4, nums, alert_status)
                        set_cell(sheet, 10, 5, other_info, alert_status)
                    elif status == "Terminating":
                        set_cell(sheet, 11, 4, nums, status)
                        set_cell(sheet, 11, 5, other_info, alert_status)
                    elif status == "ERROR":
                        set_cell(sheet, 12, 4, nums, alert_status)
                        set_cell(sheet, 12, 5, other_info, alert_status)
                    elif status == "restart":
                        set_cell(sheet, 13, 4, nums, alert_status)
                        set_cell(sheet, 13, 5, other_info, alert_status)
        else:
            set_cell(sheet, 6, 4, "无结果", "error")
            set_cell(sheet, 7, 4, "无结果", "error")
            set_cell(sheet, 8, 4, "无结果", "error")
            set_cell(sheet, 9, 4, "无结果", "error")
            set_cell(sheet, 10, 4, "无结果", "error")
            set_cell(sheet, 11, 4, "无结果", "error")
            set_cell(sheet, 12, 4, "无结果", "error")
            set_cell(sheet, 13, 4, "无结果", "error")
        if "cluster_resources" in items.keys():
            key_data = items["cluster_resources"]
            for data in key_data:
                alert_status = data["alert_status"]
                check_point = data["check_data"]["resources"]
                if check_point == "mem":
                    set_cell(sheet, 16, 4, data["check_data"]["percent"], alert_status)
                    del data["check_data"]["resources"]
                    set_cell(sheet, 16, 5, str(data["check_data"]), alert_status)
                if check_point == "cpu":
                    set_cell(sheet, 15, 4, data["check_data"]["percent"], alert_status)
                    del data["check_data"]["resources"]
                    set_cell(sheet, 15, 5, str(data["check_data"]), alert_status)
        else:
            set_cell(sheet, 15, 4, "无结果", "error")
            set_cell(sheet, 16, 4, "无结果", "error")
        if "pod_ip" in items.keys():
            key_data = items["pod_ip"]
            alert_status = key_data[0]["alert_status"]
            values = str(key_data[0]["check_data"]["percent"])
            set_cell(sheet, 17, 4, values, alert_status)
        else:
            set_cell(sheet, 17, 4, "无结果", "error")
        if "svc_ip" in items.keys():
            key_data = items["svc_ip"]
            alert_status = key_data[0]["alert_status"]
            values = str(key_data[0]["check_data"]["percent"])
            set_cell(sheet, 18, 4, values, alert_status)
        else:
            set_cell(sheet, 18, 4, "无结果", "error")
        if "coredns_replicas" in items.keys():
            key_data = items["coredns_replicas"]
            alert_status = key_data[0]["alert_status"]
            values = key_data[0]["check_data"]["readyReplicas"]
            other_info = str(key_data[0]["check_data"])
            set_cell(sheet, 20, 4, values, alert_status)
            set_cell(sheet, 20, 5, other_info, alert_status)
        else:
            set_cell(sheet, 20, 4, "无结果", "error")
        if "pod_dns" in items.keys():
            key_data = items["pod_dns"]
            alert_status = "info"
            other_info = []
            for data in key_data:
                if data["alert_status"] == "error":
                    alert_status = "error"
                    other_info.append(data["check_data"])
            set_cell(sheet, 21, 4, alert_status, alert_status)
            set_cell(sheet, 21, 5, " ".join(other_info), alert_status)
        else:
            set_cell(sheet, 21, 4, "无结果", "error")
        if "weave_status" in items.keys():
            key_data = items["weave_status"]
            other_info = []
            ok_nums = 0
            error_nums = 0
            for data in key_data:
                if data["alert_status"] == "error":
                    error_nums += 1
                    other_info.append(data["check_data"]["ip"])
                else:
                    ok_nums += 1
            if error_nums > 0:
                set_cell(sheet, 24, 4, str(error_nums))
                set_cell(sheet, 24, 5, " ".join(other_info), "error")
            set_cell(sheet, 23, 4, str(ok_nums))


def component_json_to_excel(component_dict: dict, sheet: Workbook.worksheets):
    title_list = ["apiserver_process", "apiserver_health", "controller_process", "controller_health",
                  "scheduler_process", "scheduler_health", "etcd_process", "etcd_health", "etcd_endpoint"]
    ip_list = list(component_dict.keys())
    for ip in ip_list:
        items = component_dict[ip]
        row = ip_list.index(ip) + 2
        set_cell(sheet, row, 1, ip)
        for data_key in title_list:
            column = title_list.index(data_key) + 2
            if data_key in items.keys():
                if data_key != "etcd_endpoint":
                    set_cell(sheet, row, column, items[data_key][0]["check_data"], items[data_key][0]["alert_status"])
                else:
                    alert_status = "info"
                    other_info = list()
                    for data_key_data in items[data_key]:
                        if data_key_data["alert_status"] == "error":
                            alert_status = data_key_data["alert_status"]
                            other_info.append(data_key_data["check_data"]["url"])
                    if alert_status == "info":
                        set_cell(sheet, row, column, items[data_key][0]["check_data"]["status"])
                    else:
                        set_cell(sheet, row, column, " ".join(other_info), alert_status)
            else:
                set_cell(sheet, row, column, "无结果", "error")


def node_json_to_excel(node_dict: dict, cluster_dict: dict, sheet: Workbook.worksheets):
    title_list = ["systemLoad", "node_resources_cpu", "node_resources_mem", "diskUsage", "diskIO", "nicTraffic",
                  "systemFD", "conntrack", "pidNUM", "check_node_dns", "containerdProcess", "dockerProcess",
                  "dockerIsHanged", "dockerFD", "nullPIDContainers", "kubeletProcess", "kubeletPortCheck",
                  "kubeProxyPortCheck", "weaverStatus", "weaverConnection", "weaverDiscloseIP", "ZProcess", "ntpTime"]
    a_list = ["systemLoad", "dockerProcess", "dockerIsHanged", "dockerFD", "nullPIDContainers", "containerdProcess",
              "kubeletProcess", "kubeletPortCheck", "kubeProxyPortCheck", "systemFD", "conntrack", "pidNUM",
              "weaverStatus", "node_resources_cpu", "node_resources_mem", "weaverDiscloseIP"]
    cluster_ip = list(cluster_dict.keys())[0]
    for resources_data in cluster_dict[cluster_ip]["node_resources"]:
        cluster_ip = resources_data["check_data"]["ip"]
        if cluster_ip not in node_dict.keys():
            node_dict[cluster_ip] = {}
        if resources_data["check_data"]["resources"] == "cpu":
            del resources_data["check_data"]["ip"]
            del resources_data["check_data"]["resources"]
            node_dict[cluster_ip]["node_resources_cpu"] = list()
            node_dict[cluster_ip]["node_resources_cpu"].append(resources_data)
        elif resources_data["check_data"]["resources"] == "mem":
            del resources_data["check_data"]["ip"]
            del resources_data["check_data"]["resources"]
            node_dict[cluster_ip]["node_resources_mem"] = list()
            node_dict[cluster_ip]["node_resources_mem"].append(resources_data)
    ip_list = list(node_dict.keys())
    for ip in ip_list:
        items = node_dict[ip]
        row = ip_list.index(ip) + 2
        set_cell(sheet, row, 1, ip)
        for data_key in title_list:
            column = title_list.index(data_key) + 2
            if data_key in items.keys():
                if data_key in a_list:
                    values = ""
                    if items[data_key][0]["check_data"] == "":
                        if items[data_key][0]["alert_status"] == "info":
                            values = "ok"
                        elif items[data_key][0]["alert_status"] == "error":
                            values = "failed"
                    else:
                        if data_key == "node_resources_cpu" or data_key == "node_resources_mem":
                            values = items[data_key][0]["check_data"]["percent"]
                        elif data_key == "systemFD":
                            values = items[data_key][0]["check_data"]["filePercentage"]
                        elif data_key == "conntrack":
                            values = items[data_key][0]["check_data"]["usedConntrackPercentage"]
                        elif data_key == "pidNUM":
                            values = items[data_key][0]["check_data"]["pidUsedPercentage"]
                        elif data_key == "dockerFD":
                            values = items[data_key][0]["check_data"]["dockerFDUsedPercentage"]
                        elif data_key == "nullPIDContainers":
                            values = items[data_key][0]["check_data"]["nullPIDContainerID"]
                        else:
                            values = str(items[data_key][0]["check_data"])
                    set_cell(sheet, row, column, values, items[data_key][0]["alert_status"])
                if data_key in ["diskUsage", "weaverConnection", "ZProcess", "ntpTime"]:
                    if items[data_key][0]["alert_status"] == "error":
                        try:
                            values = " ".join(items[data_key][0]["others"])
                        except KeyError:
                            values = ""
                    else:
                        if data_key == "ntpTime":
                            values = items[data_key][0]["check_data"]["timeDiff"]
                        else:
                            values = "ok"
                    set_cell(sheet, row, column, values, items[data_key][0]["alert_status"])
                if data_key in ["diskIO", "nicTraffic", "check_node_dns"]:
                    values = []
                    alert_status = "info"
                    for data in items[data_key]:
                        if data["alert_status"] == "error":
                            if data_key == "check_node_dns":
                                values.append(data["check_data"])
                            else:
                                values = data["others"]
                            alert_status = data["alert_status"]
                    if alert_status == "info":
                        values = "ok"
                    else:
                        values = " ".join(values)
                    set_cell(sheet, row, column, values, alert_status)
            else:
                set_cell(sheet, row, column, "无结果", "error")


def cert_to_excel(component_dict: dict, sheet: Workbook.worksheets):
    title_list = ["cert", "endTime", "remainingTime"]
    ip_list = list(component_dict.keys())
    row = 1
    for ip in ip_list:
        items = component_dict[ip]
        for i in range(len(items["cert_time"])):
            row += 1
            for data_key in title_list:
                column = title_list.index(data_key) + 2
                set_cell(sheet, row, 1, ip)
                set_cell(sheet, row, column, items["cert_time"][i]["check_data"][data_key],
                         items["cert_time"][i]["alert_status"])


def generate_template_excel(path: str, file_name: str, cluster: dict, component: dict, node: dict):
    path = path.rstrip("/")
    excel_file = "{0}/{1}.xlsx".format(path, file_name)
    check_output_file(excel_file)
    wb = Workbook()
    cluster_sheet = wb.active
    cluster_sheet.title = "k8s集群状态"
    cluster_sheet.merge_cells('c1:d1')
    cluster_sheet.merge_cells('a2:a24')
    cluster_sheet.merge_cells('b2:b4')
    cluster_sheet.merge_cells('b5:b13')
    cluster_sheet.merge_cells('b14:b18')
    cluster_sheet.merge_cells('b19:b21')
    cluster_sheet.merge_cells('b22:b24')
    cluster_sheet['a1'] = "集群"
    cluster_sheet['e1'] = "备注"
    title_tuple = [
        ["检查点", "检查结果", ""],
        ["节点", "状态", "数量"],
        ["", "Ready", "0"],
        ["", "NotReady", "0"],
        ["pod（kube-system）", "状态", "数量"],
        ["", "Running", "0"],
        ["", "Completed", "0"],
        ["", "CrashLoopBackOff", "0"],
        ["", "ImagePullBackOff", "0"],
        ["", "ContainerCreating", "0"],
        ["", "Terminating", "0"],
        ["", "Error", "0"],
        ["", "restart", "0"],
        ["资源", "类型", "值"],
        ["", "CPU", "0"],
        ["", " 内存", "0"],
        ["", "Pod IP", "0"],
        ["", "svc IP", "0"],
        ["DNS", "检查项", "结果"],
        ["", "CoreDNS副本数", "0"],
        ["", "Pod域名解析", "0"],
        ["weave", "状态", "数量"],
        ["", "Ok", "0"],
        ["", "Error", "0"]
    ]
    title_font = Font(u'宋体', bold=True, size=14)
    title_fill = PatternFill(patternType="solid", start_color="32CD32")
    title_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin", color="D3D3D3"),
                    right=Side(style="thin", color="D3D3D3"),
                    top=Side(style="thin", color="D3D3D3"),
                    bottom=Side(style="thin", color="D3D3D3"))
    subtitle_font = Font(u'宋体', bold=True)
    subtitle_fill = PatternFill(patternType="solid", start_color="90EE90")
    set_column_width(cluster_sheet, "A", "D", 20)
    set_column_width(cluster_sheet, "E", "E", 50)
    set_row_height(cluster_sheet, 1, 24, 20)
    set_range_styles(cluster_sheet, 'a1:e1', font=title_font, fill=title_fill, alignment=title_alignment)
    set_range_styles(cluster_sheet, 'a1:e24', alignment=title_alignment, border=border)
    set_range_styles(cluster_sheet, 'c2:d2', font=subtitle_font, fill=subtitle_fill)
    set_range_styles(cluster_sheet, 'c5:d5', font=subtitle_font, fill=subtitle_fill)
    set_range_styles(cluster_sheet, 'c14:d14', font=subtitle_font, fill=subtitle_fill)
    set_range_styles(cluster_sheet, 'c19:d19', font=subtitle_font, fill=subtitle_fill)
    set_range_styles(cluster_sheet, 'c22:d22', font=subtitle_font, fill=subtitle_fill)
    set_ranger_cell_values(cluster_sheet, 'b1:d24', title_tuple)
    component_sheet = wb.create_sheet(title="k8s组件状态", index=1)
    component_title_list = ["节点", "apiserver进程", "apiserver健康", "controller进程", "controller健康", "scheduler进程",
                            "scheduler健康", "etcd进程", "etcd健康", "etcd endpoint"]
    component_sheet.append(component_title_list)
    set_range_styles(component_sheet, 'a1:j1', font=title_font, fill=title_fill, alignment=title_alignment)
    set_column_width(component_sheet, "A", "J", 20)
    node_sheet = wb.create_sheet(title="节点信息", index=2)
    node_title_list = ["节点", "负载", "cpu request",
                       "mem request", "磁盘容量", "磁盘IO", "网卡", "文件打开数", "conntrack使用率", "线程数", "DNS解析",
                       "containerd状态", "docker进程", "docker夯住", "docker描述符", "docker容器pid校验", "kubelet进程", "kubelet健康",
                       "kube-proxy健康", "weaver状态", "weaver establish", "weave泄露IP", "Z状态进程", "NTP时间差"]
    node_sheet.append(node_title_list)
    set_range_styles(node_sheet, 'a1:x1', font=title_font, fill=title_fill, alignment=title_alignment)
    set_column_width(node_sheet, "A", "X", 20)
    cert_sheet = wb.create_sheet(title="证书信息", index=3)
    cert_tile_list = ["节点", "证书名称", "过期时间", "剩余时间(day)"]
    cert_sheet.append(cert_tile_list)
    set_range_styles(cert_sheet, 'a1:d1', font=title_font, fill=title_fill, alignment=title_alignment)
    set_column_width(cert_sheet, "A", "D", 30)
    cluster_json_to_excel(cluster, cluster_sheet)
    component_json_to_excel(component, component_sheet)
    node_json_to_excel(node, cluster, node_sheet)
    cert_to_excel(component, cert_sheet)
    wb.save(excel_file)


def format_json_to_excel(path: str, file_name: str):
    file_list = get_suffix_file(path, ".txt")
    check_data_dict = check_file_from_what(file_list)
    cluster_info = format_txt_to_json(check_data_dict["cluster"])
    component_info = format_txt_to_json(check_data_dict["component"])
    node_info = format_txt_to_json(check_data_dict["node"])
    generate_template_excel(path, file_name, cluster_info, component_info, node_info)


if __name__ == '__main__':
    parse = get_parse()
    args = parse.parse_args()
    format_json_to_excel(args.data_dir, args.cluster_name)
